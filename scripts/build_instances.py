#!/usr/bin/env python3
import csv
import json
import os
from collections import Counter, defaultdict
from datetime import datetime
from statistics import mean
from typing import Dict, Iterable, List, Tuple

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR = os.path.join(ROOT_DIR, "data", "raw")
OUT_DIR = os.path.join(ROOT_DIR, "data", "instances")

FAST_KEYWORDS = [
    "espresso", "coffee", "americano", "drip", "filter", "brew", "black", "short",
]
MEDIUM_KEYWORDS = [
    "latte", "cappuccino", "mocha", "macchiato", "flat white", "cortado",
    "hot chocolate", "chai", "matcha", "tea",
]
SLOW_KEYWORDS = [
    "sandwich", "panini", "bagel", "muffin", "croissant", "cake", "cookie",
    "scone", "waffle", "toast", "salad", "breakfast", "lunch", "meal",
    "smoothie", "frappe", "frappuccino", "iced blended", "food",
]


def classify_item(name: str) -> str:
    n = name.lower().strip()
    for kw in SLOW_KEYWORDS:
        if kw in n:
            return "slow"
    for kw in MEDIUM_KEYWORDS:
        if kw in n:
            return "medium"
    for kw in FAST_KEYWORDS:
        if kw in n:
            return "fast"
    return "medium"


def hourly_counts(datetimes: Iterable[datetime]) -> Counter:
    counts = Counter()
    for dt in datetimes:
        counts[dt.hour] += 1
    return counts


def peak_hour_and_rate(hour_counts: Counter, days: int) -> Tuple[int, float]:
    if not hour_counts:
        return 9, 0.0
    peak_hour = max(hour_counts, key=hour_counts.get)
    rate_per_hour = hour_counts[peak_hour] / max(days, 1)
    return peak_hour, rate_per_hour


def compute_mix(type_counts: Counter) -> Dict[str, float]:
    total = sum(type_counts.values())
    if total == 0:
        return {"fast": 0.5, "medium": 0.3, "slow": 0.2}
    return {
        "fast": type_counts.get("fast", 0) / total,
        "medium": type_counts.get("medium", 0) / total,
        "slow": type_counts.get("slow", 0) / total,
    }


def parse_hf() -> Dict:
    path = os.path.join(RAW_DIR, "hf_coffee_sales_index.csv")
    datetimes = []
    type_counts = Counter()
    dates = set()
    if not os.path.exists(path):
        return {}
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            dt = datetime.fromisoformat(row["datetime"])
            datetimes.append(dt)
            dates.add(dt.date())
            t = classify_item(row.get("coffee_name", ""))
            type_counts[t] += 1
    hour_counts = hourly_counts(datetimes)
    peak_hour, rate = peak_hour_and_rate(hour_counts, len(dates))
    return {
        "source": "huggingface_coffeesales",
        "peak_hour": peak_hour,
        "total_rate_per_hour": rate,
        "mix": compute_mix(type_counts),
    }


def parse_kaggle() -> Dict:
    path = os.path.join(RAW_DIR, "kaggle_coffee_sales_dataset", "Coffe_sales.csv")
    datetimes = []
    type_counts = Counter()
    dates = set()
    if not os.path.exists(path):
        return {}
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            date_str = row.get("Date") or row.get("date")
            time_str = row.get("Time") or row.get("time")
            if not date_str or not time_str:
                continue
            dt = datetime.fromisoformat(f"{date_str} {time_str}")
            datetimes.append(dt)
            dates.add(dt.date())
            t = classify_item(row.get("coffee_name", ""))
            type_counts[t] += 1
    hour_counts = hourly_counts(datetimes)
    peak_hour, rate = peak_hour_and_rate(hour_counts, len(dates))
    return {
        "source": "kaggle_coffee_sales_dataset",
        "peak_hour": peak_hour,
        "total_rate_per_hour": rate,
        "mix": compute_mix(type_counts),
    }


def parse_maven() -> Dict:
    if openpyxl is None:
        return {
            "source": "maven_coffee_shop_sales",
            "error": "openpyxl not installed",
        }
    path = os.path.join(RAW_DIR, "maven_coffee_shop_sales", "Coffee Shop Sales.xlsx")
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    # try common column names
    date_col = None
    time_col = None
    item_col = None
    for cand in ["transaction_date", "transaction_date", "date"]:
        if cand in idx:
            date_col = cand
            break
    for cand in ["transaction_time", "time", "transaction_time"]:
        if cand in idx:
            time_col = cand
            break
    for cand in ["product_name", "item_name", "product", "item"]:
        if cand in idx:
            item_col = cand
            break

    if date_col is None or time_col is None:
        # fallback: try to detect by substring
        for h in headers:
            if "date" in str(h).lower() and date_col is None:
                date_col = h
            if "time" in str(h).lower() and time_col is None:
                time_col = h
            if "product" in str(h).lower() and item_col is None:
                item_col = h

    datetimes = []
    type_counts = Counter()
    dates = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = row[idx[date_col]] if date_col in idx else None
        time_val = row[idx[time_col]] if time_col in idx else None
        item_val = row[idx[item_col]] if item_col in idx else ""
        if date_val is None or time_val is None:
            continue
        # openpyxl may return datetime/date/time objects
        if isinstance(date_val, datetime):
            date_part = date_val.date()
        else:
            date_part = date_val
        if isinstance(time_val, datetime):
            time_part = time_val.time()
        else:
            time_part = time_val
        try:
            dt = datetime.combine(date_part, time_part)
        except Exception:
            # try parsing strings
            try:
                dt = datetime.fromisoformat(f"{date_val} {time_val}")
            except Exception:
                continue
        datetimes.append(dt)
        dates.add(dt.date())
        t = classify_item(str(item_val))
        type_counts[t] += 1

    hour_counts = hourly_counts(datetimes)
    peak_hour, rate = peak_hour_and_rate(hour_counts, len(dates))
    return {
        "source": "maven_coffee_shop_sales",
        "peak_hour": peak_hour,
        "total_rate_per_hour": rate,
        "mix": compute_mix(type_counts),
    }


def build_instances(source_info: Dict, instance_prefix: str) -> List[Dict]:
    if not source_info or source_info.get("total_rate_per_hour", 0) == 0:
        return []
    base_rate = source_info["total_rate_per_hour"]
    mix = source_info["mix"]

    scenarios = []
    scales = [0.7, 0.85, 1.0, 1.15, 1.3]
    staff_options = [
        {"order_attendants": 1, "baristas": 2, "pickup_attendants": 1},
        {"order_attendants": 2, "baristas": 2, "pickup_attendants": 1},
        {"order_attendants": 1, "baristas": 3, "pickup_attendants": 1},
        {"order_attendants": 1, "baristas": 2, "pickup_attendants": 2},
    ]
    patience_options = [
        {"fast": 2.5, "medium": 5.0, "slow": 8.0},
        {"fast": 3.0, "medium": 6.0, "slow": 10.0},
        {"fast": 4.0, "medium": 8.0, "slow": 12.0},
    ]
    prep_alpha_options = [0.05, 0.10, 0.15]

    idx = 1
    for scale in scales:
        for staff in staff_options:
            for patience in patience_options:
                for alpha in prep_alpha_options:
                    arrival_rates = {
                        k: base_rate * scale * mix[k] for k in ("fast", "medium", "slow")
                    }
                    scenario = {
                        "name": f"{instance_prefix}_{idx:03d}",
                        "source": source_info["source"],
                        "peak_hour": source_info.get("peak_hour", 9),
                        "arrival_rates_per_hour": arrival_rates,
                        "order_time_mean": {"fast": 0.8, "medium": 1.2, "slow": 1.8},
                        "prep_time_mean": {"fast": 1.0, "medium": 2.5, "slow": 4.0},
                        "patience_mean": patience,
                        "staffing": staff,
                        "prep_queue_alpha": alpha,
                        "hours": 4.0,
                        "warmup_hours": 0.5,
                    }
                    scenarios.append(scenario)
                    idx += 1
    return scenarios


def main() -> None:
    os.makedirs(OUT_DIR, exist_ok=True)

    sources = []
    sources.append(parse_maven())
    sources.append(parse_hf())
    sources.append(parse_kaggle())

    all_instances = []
    for src in sources:
        if not src:
            continue
        if src.get("error"):
            print(f"Skipping {src.get('source')} - {src.get('error')}")
            continue
        prefix = src["source"].replace("-", "_")
        all_instances.extend(build_instances(src, prefix))

    for inst in all_instances:
        path = os.path.join(OUT_DIR, f"{inst['name']}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(inst, f, indent=2)
        print(f"Wrote {path}")

    summary_path = os.path.join(OUT_DIR, "summary.json")
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(all_instances, f, indent=2)
    print(f"Wrote {summary_path}")


if __name__ == "__main__":
    main()
